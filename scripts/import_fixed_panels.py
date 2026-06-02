#!/usr/bin/env python3
"""import_fixed_panels.py — convert NCKU's WES-I / WES-II / WGS panel
spreadsheets into the gene-panel format the UI already understands,
plus a JSON index so the new "Fixed panels" UI tabs can list them.

Sources (under repo's fixed_panel/):
  WES-I.xlsx                        — 6 sheets, one per 科別 (skin /
                                      eye / ENT / pediatric / neuro /
                                      oncology). Each column = a panel:
                                        row 1 = name (範本名稱)
                                        row 2 = disease label
                                        row 3 = references
                                        row 4+ = gene symbols
  WES-II.xlsx                       — same layout
  other_panel/<Spec>_<topic>.csv    — WGS panels, columns:
                                        SNV/indel | CNV | STR | Mito
                                      filename prefix groups the panel
                                      under a specialty (Neurology /
                                      Pediatric / ENT).
  other_panel/*.xlsx                — same csv-equivalent layout, just
                                      delivered as xlsx.

Per spec: merge SNV/indel + CNV + Mito columns into one deduped gene
list per panel; drop STR entirely. Each resulting panel becomes a
plain-text file (one gene per line) and lands under

  {PHENO_DATA_DIR}/fixed_panels/{series}/{category}/{panel_name}.txt

A companion JSON index at

  {PHENO_DATA_DIR}/fixed_panels/index.json

groups the panels for the UI:
  {
    "series": [
      {"key": "WES-I",  "label": "WES-I",  "groups": [
          {"category": "皮膚科", "panels": [
              {"name": "EB",  "key": "WES-I__皮膚科__EB",  "gene_count": 122},
              ...
          ]},
          ...
      ]},
      ...
    ]
  }

The panel's `key` is what gets passed to `phenotype_scorer` (it
matches a file written into the gene panel store), so the existing
phenotype-scoring code consumes them without changes.

Idempotent — wipes the fixed_panels/ output dir each run.

Usage:
    PYTHONPATH=backend python scripts/import_fixed_panels.py [--src DIR] [--out DIR]
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))
from app.config import PHENO_DATA_DIR  # noqa: E402


# ── Sheet → category routing ──────────────────────────────────────

# WES-I.xlsx + WES-II.xlsx sheet names ARE the category names (科別).
# For WGS (other_panel/), the category comes from the filename prefix.
_WGS_PREFIX_TO_CATEGORY = {
    "ENT":        "耳鼻喉科",
    "Neurology":  "神經科",
    "Pediatric":  "兒科",
    "Cardiology": "心臟科",
    "Oncology":   "腫瘤醫學",
    "Dermatology":"皮膚科",
    "Renal":      "腎臟科",
    "Endocrine":  "內分泌",
}


def _slug(s: str, max_len: int = 80) -> str:
    """Filesystem-safe key. Keep CJK + alphanumerics + a few
    separators, replace everything else with `_`."""
    s = re.sub(r"[\s/]+", "_", s.strip())
    s = re.sub(r"[^\w一-鿿\-+]+", "", s)
    return s[:max_len] or "panel"


def _is_gene_symbol(cell: str) -> bool:
    """Quick filter — anything starting with `[A-Z]`, no whitespace,
    < 30 chars is treated as a gene symbol candidate. Drops header
    rows / blanks / multi-word disease names that bled into a cell.
    """
    if not cell:
        return False
    s = cell.strip()
    if not s or " " in s or len(s) > 30:
        return False
    return bool(re.match(r"^[A-Z][A-Z0-9\-\.@]*$", s))


# ── WES-I / WES-II xlsx readers ───────────────────────────────────

_PANEL_NAME_CORRECTIONS = {
    # Typo in the source spreadsheet. Keep UI labels reviewer-friendly.
    "Syndromic hearing losss": "Syndromic hearing loss",
}


def _panel_display_name(label: str) -> str:
    """Drop the spreadsheet's specialty / tier prefix without
    truncating meaningful hyphenated panel names such as
    ``Non-syndromic hearing loss``.
    """
    display = re.sub(r"^[^-]+-(?:I|II)-", "", label.strip())
    # WES-II's pediatric metabolism panel uses a specialty-only prefix.
    display = re.sub(r"^婦兒科-", "", display)
    display = display.strip()
    return _PANEL_NAME_CORRECTIONS.get(display, display)


def _find_gene_list_row(rows: list[tuple]) -> int | None:
    """Return the first row containing the gene-panel-list marker.

    The marker row also contains the first gene in each panel column,
    so callers must include it in the data range.
    """
    for row_idx, row in enumerate(rows):
        marker = re.sub(r"\s+", "", str(row[0] if row else "")).lower()
        if "panel" in marker and "list" in marker:
            return row_idx
    return None


def _read_wes_xlsx(path: Path, series_key: str) -> list[dict]:
    """Return one record per panel column across all sheets.
    Each record: {series, category, name, key, genes: [...] }."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    out: list[dict] = []
    for sheet_name in wb.sheetnames:
        category = sheet_name.strip()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Header row: "範本名稱" + panel names per column. Some sheets
        # have the panel name in col 1 only (single-panel sheet) — skip
        # empty / header-marker columns.
        headers = list(rows[0])
        # Strip the row's leading label cell ("範本名稱" / similar) — it's
        # not a panel name itself.
        panel_cols: list[tuple[int, str, str]] = []
        for col_idx, h in enumerate(headers):
            if col_idx == 0:
                continue
            label = ("" if h is None else str(h)).strip()
            if not label:
                continue
            # Remove the known specialty / tier prefix only. Splitting
            # on the final `-` corrupts names like "Non-syndromic".
            display = _panel_display_name(label) or label
            # Keep the historical key stable because analyses persist
            # it in sample metadata. The corrected display name is a
            # separate UI concern.
            key_name = label.split("-")[-1].strip() or label
            panel_cols.append((col_idx, display, key_name))
        if not panel_cols:
            continue

        # Walk the explicit gene-list block only. Disease labels and
        # references can contain gene-looking tokens (for example
        # HPO, OMIM, or journal abbreviations), so parsing every row
        # silently inflates panel counts.
        gene_list_row = _find_gene_list_row(rows)
        if gene_list_row is None:
            continue
        per_panel_genes: dict[int, list[str]] = {col: [] for col, _, _ in panel_cols}
        for r in rows[gene_list_row:]:
            for col_idx, _, _ in panel_cols:
                if col_idx >= len(r):
                    continue
                cell = r[col_idx]
                if cell is None:
                    continue
                # Some "腫瘤醫學" sheets pack the whole gene list as a
                # comma-separated cell on a single row — split first.
                for token in re.split(r"[,;\s]+", str(cell)):
                    tok = token.strip().upper()
                    if _is_gene_symbol(tok):
                        per_panel_genes[col_idx].append(tok)

        for col_idx, name, key_name in panel_cols:
            genes = sorted(set(per_panel_genes[col_idx]))
            if not genes:
                continue
            key = f"{series_key}__{category}__{_slug(key_name)}"
            out.append({
                "series":   series_key,
                "category": category,
                "name":     name,
                "key":      key,
                "genes":    genes,
            })
    wb.close()
    return out


# ── WGS other_panel readers (csv + xlsx, all with the
# SNV/indel | CNV | STR | Mito layout) ────────────────────────────

def _row_to_genes(cells: list[str]) -> tuple[list[str], list[str], list[str]]:
    """Return (snv, cnv, mito) — STR dropped per spec."""
    snv  = [str(cells[0]).strip().upper()] if len(cells) > 0 and cells[0] else []
    cnv  = [str(cells[1]).strip().upper()] if len(cells) > 1 and cells[1] else []
    mito = [str(cells[3]).strip().upper()] if len(cells) > 3 and cells[3] else []
    return snv, cnv, mito


def _read_wgs_csv(path: Path) -> list[str]:
    snv, cnv, mito = [], [], []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        next(reader, None)  # skip header
        for row in reader:
            s, c, m = _row_to_genes(row)
            snv += s; cnv += c; mito += m
    return sorted({g for g in snv + cnv + mito if _is_gene_symbol(g)})


def _read_wgs_xlsx(path: Path) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    snv, cnv, mito = [], [], []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # First row may be a title (1 cell), second a header
        # ("SNV/indel | CNV | STR | Mitochondria"). Skip rows whose
        # first cell is not a gene-looking symbol.
        for r in rows:
            cells = [("" if c is None else str(c)) for c in (r or [])]
            if cells and _is_gene_symbol(cells[0].strip().upper()):
                s, c, m = _row_to_genes(cells)
                snv += s; cnv += c; mito += m
    wb.close()
    return sorted({g for g in snv + cnv + mito if _is_gene_symbol(g)})


def _read_wgs_panels(other_dir: Path) -> list[dict]:
    """Iterate every csv + xlsx under other_panel/ and emit one
    record per file. Category derived from filename prefix."""
    out: list[dict] = []
    files = sorted(list(other_dir.glob("*.csv")) + list(other_dir.glob("*.xlsx")))
    for f in files:
        stem = f.stem
        # "Neurology_ALS_HSP" → category "神經科" + name "ALS_HSP"
        m = re.match(r"^([A-Za-z]+)[_\-](.+)$", stem)
        prefix, rest = (m.group(1), m.group(2)) if m else ("", stem)
        category = _WGS_PREFIX_TO_CATEGORY.get(prefix, prefix or "其他")
        name = rest.replace("_", " ")
        genes = _read_wgs_xlsx(f) if f.suffix.lower() == ".xlsx" else _read_wgs_csv(f)
        if not genes:
            continue
        key = f"WGS__{category}__{_slug(name)}"
        out.append({
            "series":   "WGS",
            "category": category,
            "name":     name,
            "key":      key,
            "genes":    genes,
        })
    return out


# ── Output ────────────────────────────────────────────────────────

def _write_panels(records: list[dict], out_root: Path) -> None:
    """Write one .txt per panel + index.json + 1 gene-per-line files
    inside the existing GENE_PANELS_DIR so phenotype_scorer can load
    them without any code change."""
    from app.config import GENE_PANELS_DIR

    keys = [rec["key"] for rec in records]
    if len(keys) != len(set(keys)):
        duplicates = sorted({key for key in keys if keys.count(key) > 1})
        raise ValueError(f"duplicate fixed-panel keys: {', '.join(duplicates)}")

    if out_root.is_dir():
        shutil.rmtree(out_root)
    out_root.mkdir(parents=True, exist_ok=True)

    # Per-series-per-category dir tree + index entries.
    index: dict[str, dict] = {}
    for rec in records:
        sd = out_root / rec["series"] / _slug(rec["category"])
        sd.mkdir(parents=True, exist_ok=True)
        (sd / f"{_slug(rec['name'])}.txt").write_text(
            "\n".join(rec["genes"]) + "\n", encoding="utf-8")
        index.setdefault(rec["series"], {}).setdefault(rec["category"], []).append({
            "name":       rec["name"],
            "key":        rec["key"],
            "gene_count": len(rec["genes"]),
        })

    # Plus: drop a panel file in GENE_PANELS_DIR so phenotype_scorer's
    # existing loader picks them up — the panel name = the unique key.
    # Clear stale series files first so renamed/removed panels disappear.
    GENE_PANELS_DIR.mkdir(parents=True, exist_ok=True)
    for prefix in ("WES-I__", "WES-II__", "WGS__"):
        for f in GENE_PANELS_DIR.glob(f"{prefix}*.txt"):
            try: f.unlink()
            except OSError: pass
    for rec in records:
        (GENE_PANELS_DIR / f"{rec['key']}.txt").write_text(
            "\n".join(rec["genes"]) + "\n", encoding="utf-8")

    # Build the consumer-facing index JSON.
    series_order = ["WES-I", "WES-II", "WGS"]
    out = {"series": []}
    for skey in series_order:
        if skey not in index:
            continue
        groups = []
        for cat in sorted(index[skey].keys()):
            panels = sorted(index[skey][cat], key=lambda p: p["name"])
            groups.append({"category": cat, "panels": panels})
        out["series"].append({"key": skey, "label": skey, "groups": groups})
    (out_root / "index.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--src", default=str(Path(__file__).resolve().parent.parent / "fixed_panel"),
                    help="repo's fixed_panel/ dir (default: ../fixed_panel from this script)")
    ap.add_argument("--out", default=str(PHENO_DATA_DIR / "fixed_panels"),
                    help="output dir for the panel txt + index.json (default: phenotype_data/fixed_panels/)")
    args = ap.parse_args()

    src = Path(args.src)
    out = Path(args.out)

    records: list[dict] = []
    for series, fn in (("WES-I", "WES-I.xlsx"), ("WES-II", "WES-II.xlsx")):
        p = src / fn
        if p.is_file():
            records += _read_wes_xlsx(p, series)
            print(f"  + {p}: {sum(1 for r in records if r['series']==series)} panels")
        else:
            print(f"  ! {p} missing — skipped", file=sys.stderr)
    other = src / "other_panel"
    if other.is_dir():
        wgs = _read_wgs_panels(other)
        records += wgs
        print(f"  + {other}: {len(wgs)} WGS panels")

    if not records:
        print("no panels found.", file=sys.stderr)
        return 1

    _write_panels(records, out)
    total = sum(len(r["genes"]) for r in records)
    print(f"\nwrote {len(records)} panels ({total} gene rows) → {out}")
    print(f"index → {out}/index.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
