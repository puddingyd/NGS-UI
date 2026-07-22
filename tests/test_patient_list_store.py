from openpyxl import Workbook

from app.services import patient_list_store


def _save_workbook(tmp_path, workbook, name="patient-list.xlsx"):
    path = tmp_path / name
    workbook.save(path)
    return path


def test_parse_xlsx_accepts_offset_headers_aliases_and_all_sheets(
    tmp_path,
    monkeypatch,
):
    wb = Workbook()
    self_pay = wb.active
    self_pay.title = "一般自費"
    self_pay.append(["NGS-WES（自費）"])
    self_pay.append([
        "簽收時間", "檢驗項目", "檢體編號", "檢驗單號", "病歷號",
        "姓名", "性別", "生日", "檢體類別", "開單醫師", "檢驗套組",
    ])
    self_pay.append([
        "2026-07-01 09:30", "NGS-WES-II (Blood)", "8BB126WE0001",
        "ORDER-1", "MRN-1", "測試甲", "女", "2000-01-01", "血液",
        "醫師甲", "非特定",
    ])

    subsidy = wb.create_sheet("罕病基金會補助")
    subsidy.append(["罕病基金會補助"])
    subsidy.append([
        "簽收時間", "檢驗項目", "檢體編號", "檢驗單號", "病歷號",
        "姓名", "性別", "生日", "檢體類別", "開單醫師", "檢驗套組",
    ])
    # Duplicate LIS ID: the first worksheet remains authoritative.
    subsidy.append([
        "2026-07-02 10:00", "NGS-WES-II (Blood)", "8BB126WE0001",
        "ORDER-2", "MRN-CHANGED", "重複列", "女", "2000-01-01", "血液",
        "醫師乙", "非特定",
    ])
    subsidy.append([
        "2026-07-03 11:00", "NGS-WGS (Blood)", "8BB126WG0002",
        "ORDER-3", "MRN-2", "測試乙", "男", "1990-01-01", "血液",
        "醫師丙", "非特定",
    ])

    source = _save_workbook(tmp_path, wb)
    rows = patient_list_store.parse_xlsx(source)

    assert [row["lis_id"] for row in rows] == ["26WE0001", "26WG0002"]
    assert rows[0] == {
        "lis_id": "26WE0001",
        "specimen": "8BB126WE0001",
        "mrn": "MRN-1",
        "name": "測試甲",
        "test_name": "NGS-WES-II (Blood)",
        "test_type": "WES",
        "department": "",
        "physician": "醫師甲",
        "sign_received_at": "2026-07-01 09:30",
    }
    assert rows[1]["test_name"] == "NGS-WGS (Blood)"
    assert rows[1]["test_type"] == "WGS"

    # Exercise the same archive + roster merge path used by POST
    # /api/patient_list, not only the parser helper.
    patient_list_dir = tmp_path / "patient_list"
    monkeypatch.setattr(patient_list_store, "PATIENT_LIST_DIR", patient_list_dir)
    result = patient_list_store.ingest_xlsx(source.read_bytes(), "test4.xlsx")

    assert result["parsed"] == 2
    assert result["added"] == 2
    assert set(patient_list_store.load_roster()) == {"26WE0001", "26WG0002"}


def test_parse_xlsx_keeps_legacy_first_column_format(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["未完成報告清單"])
    ws.append([])
    ws.append([
        "檢體編號", "病歷號", "姓名", "檢驗名稱", "科別", "開單醫師",
        "簽收時間",
    ])
    ws.append([
        "8BB126WE0092", "MRN-92", "測試個案", "NGS-WES-II", "小兒科",
        "醫師丁", "2026-07-04 08:00",
    ])

    rows = patient_list_store.parse_xlsx(
        _save_workbook(tmp_path, wb, "legacy.xlsx")
    )

    assert len(rows) == 1
    assert rows[0]["lis_id"] == "26WE0092"
    assert rows[0]["test_name"] == "NGS-WES-II"
    assert rows[0]["department"] == "小兒科"
